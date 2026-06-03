"""
Kenya payroll run export to Excel (staff payroll items).
"""
from __future__ import annotations

from decimal import Decimal
from io import BytesIO

from app.extensions import db
from app.models.payroll import PayrollItem, PayrollRun
from sqlalchemy.orm import joinedload

TWO_DP = Decimal('0.01')

# (header label, row dict key) — order defines Excel columns.
KENYA_EXPORT_COLUMNS = [
    ('Employee Name', 'employee_name'),
    ('Basic Salary', 'basic_salary'),
    ('Benefits', 'benefits'),
    ('Gross Pay', 'gross_pay'),
    ('Taxable Pay', 'taxable_pay'),
    ('SHIF', 'shif'),
    ('Total NSSF', 'total_nssf'),
    ('Welfare Kit', 'welfare_kit'),
    ('SHELLOYEES SACCO', 'shelloyees_sacco'),
    ('MAISHA BORA SACCO', 'maisha_bora_sacco'),
    ('Voluntary Pension', 'voluntary_pension'),
    ('PAYE', 'paye'),
    ('total_deductions', 'total_deductions'),
    ('NET PAY', 'net_pay'),
]

KENYA_EXPORT_HEADERS = [label for label, _ in KENYA_EXPORT_COLUMNS]

NUMERIC_KEYS = tuple(key for _, key in KENYA_EXPORT_COLUMNS if key != 'employee_name')

# Deduction line codes excluded when matching recurring/other columns by name.
_STATUTORY_CODES = frozenset({
    'NSSF', 'SHIF', 'HOUSING_LEVY', 'PAYE',
    'PENSION_PERCENT', 'PENSION_FIXED',
})


def _decimal(value) -> Decimal:
    try:
        return Decimal(str(value or 0)).quantize(TWO_DP)
    except Exception:
        return Decimal('0')


def _normalize_name(name: str) -> str:
    return ' '.join((name or '').upper().split())


def _name_matches(name: str, *required_parts: str) -> bool:
    n = _normalize_name(name)
    return all(part.upper() in n for part in required_parts)


def basic_salary_total(item: PayrollItem) -> Decimal:
    """Basic salary from earnings breakdown (BASIC line)."""
    for row in item.earnings_breakdown or []:
        if str(row.get('code') or '').upper() == 'BASIC':
            return _decimal(row.get('amount'))
    return Decimal('0')


def benefits_total(item: PayrollItem) -> Decimal:
    """Sum employee benefit earnings (BEN-* lines) for the period."""
    total = Decimal('0')
    for row in item.earnings_breakdown or []:
        code = str(row.get('code') or '').upper()
        if not code.startswith('BEN-'):
            continue
        total += _decimal(row.get('amount'))
    return total.quantize(TWO_DP)


def _deduction_lines(item: PayrollItem) -> list[dict]:
    return list(item.deductions_breakdown or [])


def _is_statutory_code(code: str) -> bool:
    c = (code or '').upper()
    if c in _STATUTORY_CODES:
        return True
    return c.startswith('NSSF')


def voluntary_pension_total(item: PayrollItem) -> Decimal:
    total = Decimal('0')
    for row in _deduction_lines(item):
        code = (row.get('code') or '').upper()
        if code in ('PENSION_PERCENT', 'PENSION_FIXED'):
            total += _decimal(row.get('amount'))
            continue
        name = row.get('name') or ''
        if _name_matches(name, 'VOLUNTARY', 'PENSION') or _normalize_name(name) == 'VOLUNTARY PENSION':
            total += _decimal(row.get('amount'))
    return total.quantize(TWO_DP)


def _named_deduction_total(item: PayrollItem, *name_parts: str) -> Decimal:
    total = Decimal('0')
    for row in _deduction_lines(item):
        code = (row.get('code') or '').upper()
        if _is_statutory_code(code):
            continue
        if code in ('PENSION_PERCENT', 'PENSION_FIXED'):
            continue
        name = row.get('name') or ''
        if _name_matches(name, *name_parts):
            total += _decimal(row.get('amount'))
    return total.quantize(TWO_DP)


def total_deductions(item: PayrollItem) -> Decimal:
    gross = _decimal(item.gross_pay)
    net = _decimal(item.net_pay)
    return (gross - net).quantize(TWO_DP)


def nssf_employee_employer_total(employee_nssf_total: Decimal) -> Decimal:
    """Employer matches employee NSSF contribution (2× employee total)."""
    return (employee_nssf_total * 2).quantize(TWO_DP)


def kenya_export_row(item: PayrollItem) -> dict:
    emp = item.employee
    return {
        'employee_name': emp.full_name if emp else f'Employee #{item.employee_id}',
        'basic_salary': basic_salary_total(item),
        'benefits': benefits_total(item),
        'gross_pay': _decimal(item.gross_pay),
        'taxable_pay': _decimal(item.taxable_pay),
        'shif': _decimal(item.shif),
        'total_nssf': _decimal(item.nssf_employee),
        'welfare_kit': _named_deduction_total(item, 'WELFARE', 'KIT'),
        'shelloyees_sacco': _named_deduction_total(item, 'SHELLOYEES', 'SACCO'),
        'maisha_bora_sacco': _named_deduction_total(item, 'MAISHA', 'BORA'),
        'voluntary_pension': voluntary_pension_total(item),
        'paye': _decimal(item.paye),
        'total_deductions': total_deductions(item),
        'net_pay': _decimal(item.net_pay),
    }


def _row_to_excel_cells(row: dict) -> list:
    cells = []
    for _, key in KENYA_EXPORT_COLUMNS:
        if key == 'employee_name':
            cells.append(row[key])
        else:
            cells.append(float(row[key]))
    return cells


def _analysis_row(label: str, totals: dict[str, Decimal], *, nssf_emp_employer: Decimal | None = None) -> list:
    """Build an analysis row; optional NSSF (Employee + Employer) only in Total NSSF column."""
    cells = []
    for _, key in KENYA_EXPORT_COLUMNS:
        if key == 'employee_name':
            cells.append(label)
        elif key == 'total_nssf' and nssf_emp_employer is not None:
            cells.append(float(nssf_emp_employer))
        elif nssf_emp_employer is not None:
            cells.append('')
        elif key in totals:
            cells.append(float(totals[key]))
        else:
            cells.append('')
    return cells


def fetch_kenya_payroll_items(run_id: int, company_id: int) -> list[PayrollItem]:
    return (
        db.session.query(PayrollItem)
        .join(PayrollRun, PayrollRun.id == PayrollItem.payroll_run_id)
        .options(joinedload(PayrollItem.employee))
        .filter(
            PayrollItem.payroll_run_id == run_id,
            PayrollRun.company_id == company_id,
        )
        .order_by(PayrollItem.employee_id)
        .all()
    )


def build_kenya_payroll_workbook(run: PayrollRun, items: list[PayrollItem]) -> BytesIO:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = f"Payroll {run.pay_year}-{run.pay_month:02d}"

    ws.append(KENYA_EXPORT_HEADERS)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    totals = {k: Decimal('0') for k in NUMERIC_KEYS}

    for item in items:
        row = kenya_export_row(item)
        ws.append(_row_to_excel_cells(row))
        for k in NUMERIC_KEYS:
            totals[k] += row[k]

    # Blank row then analysis block
    ws.append([])
    analysis_header_row = ws.max_row + 1
    ws.append(['ANALYSIS'] + [''] * (len(KENYA_EXPORT_HEADERS) - 1))
    for cell in ws[analysis_header_row]:
        cell.font = Font(bold=True, size=12)

    ws.append(_analysis_row('Total (all employees)', totals))
    nssf_combined = nssf_employee_employer_total(totals['total_nssf'])
    ws.append(
        _analysis_row(
            'NSSF (Employee + Employer)',
            totals,
            nssf_emp_employer=nssf_combined,
        )
    )

    for row_idx in range(analysis_header_row + 1, ws.max_row + 1):
        for cell in ws[row_idx]:
            cell.font = Font(bold=True)

    ws.freeze_panes = 'A2'
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 2, 40)

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return out
